"""
Cleaner for SCT Financial Performance Report.

Only processes "source" sheets (those that hold original/typed-in data),
skipping any sheet that is mostly derived from other sheets via formulas.
This avoids #REF! errors from broken cross-sheet references in the
auto-aggregated Summary/Revenue/Expenses report sheets.

Source sheets handled:
- Revenue Part: two-row header (period banner + BUDGET/ACTUAL/VARIANCE/%).
- Expenditure Part: two-row header (period banner + BUDGET/ACTUAL/VARIANCE).
- Revenue and Expenses Details: single-row header on row 3, no segment column.
- Income Segregation: skipped by default (schema not yet handled).
"""

import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook

EXCEL_FILE = "/home/prasanga/projects/mcp/data/6. SCT_Financial Performance Report_Upto Poush End, 2082.xlsx"

# A sheet is considered "derived" when at least this fraction of its
# non-empty cells are formulas referencing OTHER sheets. The four
# auto-generated report sheets in this workbook fall at ≥9.5%; the source
# sheets (Revenue Part, Expenditure Part, Revenue and Expenses Details)
# all fall at ≤6.4%. 0.07 sits cleanly in that gap.
DERIVED_THRESHOLD = 0.07

HEADER_ANCHORS = {"S.N.", "S.N", "SN", "S.NO."}

# Sheets we always skip regardless of content.
ALWAYS_SKIP = {"Sheet4", "Sheet2"}


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def clean_token(s):
    """Normalize a single string token for use in a column name."""
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return ""
    s = str(s).strip()
    if not s or s.lower() == "nan":
        return ""
    # Handle '%age' before the generic % replacement.
    s = re.sub(r"%\s*age", "PERCENTAGE", s, flags=re.IGNORECASE)
    s = (
        s.upper()
        .replace("%", "PCT")
        .replace(" ", "_")
        .replace("-", "_")
        .replace("/", "_")
        .replace("(", "")
        .replace(")", "")
        .replace(".", "")
        .replace(",", "")
    )
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def dedup_columns(names):
    seen = {}
    out = []
    for n in names:
        if not n:
            n = "UNNAMED"
        if n in seen:
            seen[n] += 1
            out.append(f"{n}_{seen[n]}")
        else:
            seen[n] = 0
            out.append(n)
    return out


def find_header_row(ws, max_scan=15):
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        if not row or row[0] is None:
            continue
        first = str(row[0]).strip().upper()
        if first in HEADER_ANCHORS:
            return row_idx
    return None


def coerce_numerics(df, keep_text_cols=()):
    keep = set(keep_text_cols)
    for col in df.columns:
        if df[col].dtype != "object" or col in keep:
            continue
        converted = pd.to_numeric(df[col], errors="coerce")
        orig_non_null = df[col].notna() & (df[col].astype(str).str.strip() != "")
        if orig_non_null.sum() == 0:
            continue
        good = converted.notna() & orig_non_null
        if good.sum() >= orig_non_null.sum() * 0.8:
            df[col] = converted
    return df


def replace_excel_errors(df):
    pattern = r"#REF!|#DIV/0!|#VALUE!|#N/A|#NAME\?|#NULL!|#NUM!"
    n = 0
    for col in df.columns:
        if df[col].dtype == "object":
            mask = df[col].astype(str).str.contains(pattern, regex=True, na=False)
            n += int(mask.sum())
            df.loc[mask, col] = np.nan
    return df, n


# ---------------------------------------------------------------------------
# Generic handler (Summary / Revenue / Expenses sheets)
# ---------------------------------------------------------------------------

def build_generic_header(ws, header_row):
    banner = (
        list(ws.iter_rows(min_row=header_row - 1, max_row=header_row - 1,
                          values_only=True))[0]
        if header_row > 1 else ()
    )
    sub = list(ws.iter_rows(min_row=header_row, max_row=header_row,
                            values_only=True))[0]

    filled = []
    cur = None
    for c in banner:
        if c is not None and str(c).strip():
            cur = str(c).strip()
        filled.append(cur)
    while len(filled) < len(sub):
        filled.append(None)

    columns = []
    period_labels = {"BUDGET", "ACTUAL", "VARIANCE"}
    for i, s in enumerate(sub):
        b = filled[i]
        b_clean = clean_token(b)
        s_clean = clean_token(s)
        if s_clean in period_labels:
            name = f"{b_clean}_{s_clean}" if b_clean else s_clean
        elif s_clean:
            name = s_clean
        elif b_clean:
            name = b_clean
        else:
            name = ""
        columns.append(name)
    return dedup_columns(columns)


def clean_generic_sheet(file_path, sheet_name):
    info = {"sheet": sheet_name, "status": "ok", "notes": []}
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    hr = find_header_row(ws)
    if hr is None:
        info["status"] = "no_header"
        return pd.DataFrame(), info
    info["header_row"] = hr

    columns = build_generic_header(ws, hr)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None,
                       skiprows=hr)
    df = df.iloc[:, :len(columns)]
    df.columns = columns

    df = df.loc[:, [c for c in df.columns if c and not c.startswith("UNNAMED")]]
    df = df.dropna(axis=1, how="all")
    df = df.dropna(how="all").reset_index(drop=True)

    df, n_err = replace_excel_errors(df)
    if n_err:
        info["notes"].append(f"Cleared {n_err} formula error cell(s).")

    seg_col = next((c for c in ("PARTICULARS", "SEGEMENT", "SEGMENT")
                    if c in df.columns),
                   df.columns[1] if len(df.columns) > 1 else None)
    if seg_col is not None:
        pat = re.compile(r"(SEGMENT|SECTION)\s*$", re.IGNORECASE)
        sections, cur = [], None
        for v in df[seg_col]:
            s = str(v).strip() if pd.notna(v) else ""
            if s and pat.search(s):
                cur = s
            sections.append(cur)
        df["SECTION_NAME"] = sections

    df = coerce_numerics(df, keep_text_cols={"SN", "SEGEMENT", "SEGMENT",
                                             "PARTICULARS", "SECTION_NAME"})

    info["shape"] = df.shape
    info["columns"] = df.columns.tolist()
    return df, info


# ---------------------------------------------------------------------------
# Revenue Part / Expenditure Part
# ---------------------------------------------------------------------------

def clean_part_sheet(file_path, sheet_name):
    info = {"sheet": sheet_name, "status": "ok", "notes": []}

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    banner_row = 4
    sub_row = 5
    data_start_row = 6

    info["header_row"] = sub_row

    # -------------------------------------------------------------------
    # READ HEADER ROWS DIRECTLY FROM OPENPYXL
    # -------------------------------------------------------------------

    banner = [
        cell for cell in next(
            ws.iter_rows(
                min_row=banner_row,
                max_row=banner_row,
                values_only=True
            )
        )
    ]

    sub = [
        cell for cell in next(
            ws.iter_rows(
                min_row=sub_row,
                max_row=sub_row,
                values_only=True
            )
        )
    ]

    max_cols = max(len(banner), len(sub))

    period_labels = {
        "BUDGET",
        "ACTUAL",
        "VARIANCE",
        "% OF ACHIEVEMENT",
    }

    columns = []

    current_banner = None

    for i in range(max_cols):

        b = banner[i] if i < len(banner) else None
        s = sub[i] if i < len(sub) else None

        b_str = str(b).strip() if b is not None and str(b).strip() else None
        s_str = str(s).strip() if s is not None and str(s).strip() else None

        if s_str and s_str.upper() in period_labels:

            if b_str:
                current_banner = b_str

            col_name = (
                f"{clean_token(current_banner)}_{clean_token(s_str)}"
                if current_banner
                else clean_token(s_str)
            )

        elif s_str:

            col_name = clean_token(s_str)
            current_banner = None

        elif b_str:

            col_name = clean_token(b_str)
            current_banner = None

        else:
            col_name = ""

        columns.append(col_name)

    columns = dedup_columns(columns)

    # -------------------------------------------------------------------
    # READ DATA ROWS DIRECTLY FROM OPENPYXL
    # -------------------------------------------------------------------

    records = []

    for row in ws.iter_rows(
        min_row=data_start_row,
        values_only=True
    ):

        # preserve exact column count
        row_values = list(row[:len(columns)])

        # skip fully empty rows
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        records.append(row_values)

    df = pd.DataFrame(records, columns=columns)

    # -------------------------------------------------------------------
    # DROP EMPTY/UNNAMED COLS
    # -------------------------------------------------------------------

    df = df.loc[
        :,
        [c for c in df.columns if c and not c.startswith("UNNAMED")]
    ]

    df = df.dropna(axis=1, how="all")
    df = df.reset_index(drop=True)

    # -------------------------------------------------------------------
    # REPLACE EXCEL ERRORS
    # -------------------------------------------------------------------

    df, n_err = replace_excel_errors(df)

    if n_err:
        info["notes"].append(
            f"Cleared {n_err} formula error cell(s)."
        )

    # -------------------------------------------------------------------
    # SEGMENT FORWARD FILL
    # -------------------------------------------------------------------

    if "SEGEMENT" in df.columns:
        df["SEGEMENT"] = df["SEGEMENT"].ffill()

    # -------------------------------------------------------------------
    # SUB SEGMENTS
    # -------------------------------------------------------------------

    if "PARTICULARS" in df.columns:

        sub_seg_pat = re.compile(r"^[IVX]+\.\s", re.IGNORECASE)

        sub_segs = []

        current_sub = None
        prev_seg = None

        for seg, part in zip(
            df.get("SEGEMENT", [None] * len(df)),
            df["PARTICULARS"]
        ):

            if seg != prev_seg:
                current_sub = None
                prev_seg = seg

            s = str(part).strip() if pd.notna(part) else ""

            if s and sub_seg_pat.match(s):
                current_sub = s

            sub_segs.append(current_sub)

        df["SUB_SEGMENT"] = sub_segs

    # -------------------------------------------------------------------
    # ROW TYPES
    # -------------------------------------------------------------------

    def classify(row):

        part = (
            str(row.get("PARTICULARS")).strip()
            if pd.notna(row.get("PARTICULARS"))
            else ""
        )

        if not part:
            return "EMPTY"

        if re.match(r"^[IVX]+\.\s", part, re.IGNORECASE):
            return "SUB_SEGMENT_HEADER"

        if part.lower().startswith("total"):
            return "TOTAL"

        return "LINE"

    if "PARTICULARS" in df.columns:
        df["ROW_TYPE"] = df.apply(classify, axis=1)

    # -------------------------------------------------------------------
    # NUMERIC COERCION
    # -------------------------------------------------------------------

    df = coerce_numerics(
        df,
        keep_text_cols={
            "SN",
            "SEGEMENT",
            "SEGMENT",
            "PARTICULARS",
            "SUB_SEGMENT",
            "ROW_TYPE",
        }
    )

    info["shape"] = df.shape
    info["columns"] = df.columns.tolist()

    return df, info

# ---------------------------------------------------------------------------
# Revenue and Expenses Details
# ---------------------------------------------------------------------------

def clean_details_sheet(file_path, sheet_name):
    info = {"sheet": sheet_name, "status": "ok", "notes": []}
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    header_row = 3
    info["header_row"] = header_row

    header = list(ws.iter_rows(min_row=header_row, max_row=header_row,
                                values_only=True))[0]

    # The header row ALSO contains the first category in cols A/B
    # (col A = '1', col B = 'Transactional Revenue'). Capture that so we
    # can seed the CATEGORY column for the rows that follow.
    first_category = None
    if (header[0] is not None and header[1] is not None
            and str(header[1]).strip()):
        first_category = str(header[1]).strip()

    columns = []
    for i, h in enumerate(header):
        if i == 0:
            columns.append("SN")
        elif i == 1:
            columns.append("LINE_ITEM")
        elif h is None or not str(h).strip():
            columns.append("")
        else:
            columns.append(clean_token(h))
    columns = dedup_columns(columns)

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None,
                       skiprows=header_row)
    df = df.iloc[:, :len(columns)]
    df.columns = columns

    df = df.loc[:, [c for c in df.columns if c and not c.startswith("UNNAMED")]]
    df = df.dropna(axis=1, how="all")
    df = df.dropna(how="all").reset_index(drop=True)

    df, n_err = replace_excel_errors(df)
    if n_err:
        info["notes"].append(f"Cleared {n_err} formula error cell(s).")

    # CATEGORY: rows where SN is a number and LINE_ITEM has text.
    # Seed with first_category captured from the header row.
    category = []
    cur_cat = first_category
    for _, row in df.iterrows():
        sn_val = row.get("SN")
        line_item = (str(row.get("LINE_ITEM")).strip()
                     if pd.notna(row.get("LINE_ITEM")) else "")
        if pd.notna(sn_val) and line_item:
            cur_cat = line_item
        category.append(cur_cat)
    df["CATEGORY"] = category

    def classify(row):
        sn_val = row.get("SN")
        line_item = (str(row.get("LINE_ITEM")).strip()
                     if pd.notna(row.get("LINE_ITEM")) else "")
        if pd.notna(sn_val) and line_item:
            return "CATEGORY"
        if line_item.lower() == "total":
            return "TOTAL"
        if line_item:
            return "LINE"
        return "EMPTY"
    df["ROW_TYPE"] = df.apply(classify, axis=1)

    df = coerce_numerics(df, keep_text_cols={"SN", "LINE_ITEM", "CATEGORY",
                                             "ROW_TYPE"})

    info["shape"] = df.shape
    info["columns"] = df.columns.tolist()
    info["notes"].append(
        "Row types: " + ", ".join(
            f"{k}={v}" for k, v in df["ROW_TYPE"].value_counts().items()
        )
    )
    return df, info


# ---------------------------------------------------------------------------
# Source vs derived sheet detection
# ---------------------------------------------------------------------------

def classify_sheets(file_path):
    """
    Inspect formulas in every sheet and decide which are SOURCE (mostly raw
    data) vs DERIVED (mostly pulled from other sheets). Returns a dict:

        { sheet_name: {
            'kind': 'source' | 'derived' | 'empty',
            'cells': int,
            'formulas': int,
            'cross_sheet_refs': int,
            'ref_errors': int,
            'references': set[str],
        } }

    Detection rule: a sheet is DERIVED when cross_sheet_refs / cells exceeds
    DERIVED_THRESHOLD. The threshold is intentionally low because even a few
    cross-sheet formulas usually mean the whole sheet is a report view.
    """
    wb = load_workbook(file_path, data_only=False)
    ref_pat = re.compile(r"'([^']+)'!|([A-Za-z_][A-Za-z0-9_ ]*)!")
    result = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        n_cells = 0
        n_formulas = 0
        n_cross = 0
        n_ref_errs = 0
        refs = set()

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                n_cells += 1
                v = str(cell.value)
                if v.startswith("="):
                    n_formulas += 1
                    for a, b in ref_pat.findall(v):
                        name = a or b
                        if name and name != sn:
                            n_cross += 1
                            refs.add(name)
                if "#REF!" in v:
                    n_ref_errs += 1

        if n_cells == 0:
            kind = "empty"
        elif n_cells > 0 and (n_cross / n_cells) > DERIVED_THRESHOLD:
            kind = "derived"
        else:
            kind = "source"

        result[sn] = {
            "kind": kind,
            "cells": n_cells,
            "formulas": n_formulas,
            "cross_sheet_refs": n_cross,
            "ref_errors": n_ref_errs,
            "references": refs,
        }

    return result


def clean_sheet(file_path, sheet_name):

    if sheet_name in ALWAYS_SKIP:
        return None, {
            "sheet": sheet_name,
            "status": "skipped",
            "notes": ["Known empty/scratch sheet."]
        }

    if sheet_name in {"Revenue Part", "Expenditure Part"}:
        return clean_part_sheet(file_path, sheet_name)

    if sheet_name == "Revenue and Expenses Details":
        return clean_details_sheet(file_path, sheet_name)

    if sheet_name == "Income Segregation":
        return None, {
            "sheet": sheet_name,
            "status": "unsupported",
            "notes": ["Layout not yet handled — different schema."]
        }

    return clean_generic_sheet(file_path, sheet_name)


def main():
    # Step 1: classify every sheet so we know which are source vs derived.
    classifications = classify_sheets(EXCEL_FILE)

    print("=" * 80)
    print("SHEET CLASSIFICATION")
    print("=" * 80)
    print(f"{'Sheet':<40} {'Kind':<10} {'Cells':>6} {'X-refs':>7} {'#REF!':>6}")
    print("-" * 80)

    for sn, info in classifications.items():
        print(
            f"{sn:<40} "
            f"{info['kind']:<10} "
            f"{info['cells']:>6} "
            f"{info['cross_sheet_refs']:>7} "
            f"{info['ref_errors']:>6}"
        )

    # Step 2: clean only the source sheets.
    cleaned = {}

    print("\n" + "=" * 80)
    print("CLEANING SOURCE SHEETS")
    print("=" * 80)

    for sheet, meta in classifications.items():

        if meta["kind"] != "source":
            continue

        if sheet in ALWAYS_SKIP:
            continue

        print("\n" + "=" * 80)
        print(f"SHEET: {sheet}")
        print("=" * 80)

        df, info = clean_sheet(EXCEL_FILE, sheet)

        if df is None:
            print(f"  Status: {info['status']}")

            for note in info["notes"]:
                print(f"  Note: {note}")

            continue

        print(f"  Header row: {info.get('header_row', '?')}")
        print(f"  Shape: {info['shape']}")

        for note in info["notes"]:
            print(f"  Note: {note}")

        print(f"  Columns: {len(info['columns'])}")

        print("\n  First 5 rows:")

        with pd.option_context(
            "display.max_columns", 8,
            "display.width", 200
        ):
            print(df.head().to_string())

        cleaned[sheet] = df

        # ------------------------------------------------------------------
        # DEBUG: Revenue Part → Card Sales value
        # ------------------------------------------------------------------
        if sheet == "Revenue Part":

            print("\n" + "=" * 80)
            print("DEBUG CARD SALES")
            print("=" * 80)

            print("\nColumns:")
            print(df.columns.tolist())

            if (
                "PARTICULARS" in df.columns
                and "MANGSIR_2082_BUDGET" in df.columns
            ):

                debug_rows = df.loc[
                    df["PARTICULARS"].astype(str).str.strip() == "Card Sales",
                    ["PARTICULARS", "MANGSIR_2082_BUDGET"]
                ]

                print("\nCard Sales rows:")
                print(debug_rows.head())

                print("\nRaw value:")
                print(debug_rows["MANGSIR_2082_BUDGET"].tolist())

            else:
                print("\nRequired columns not found.")
                print("Available columns:")
                print(df.columns.tolist())

    skipped = [
        sn
        for sn, m in classifications.items()
        if m["kind"] == "derived"
    ]

    if skipped:
        print(
            f"\n[INFO] Skipped {len(skipped)} derived sheet(s) "
            f"(would contain stale or broken references): "
            f"{', '.join(skipped)}"
        )

    print("\n" + "=" * 80)
    print(f"DONE. Cleaned {len(cleaned)} source sheet(s).")
    print("=" * 80)

    return cleaned


if __name__ == "__main__":
    main()